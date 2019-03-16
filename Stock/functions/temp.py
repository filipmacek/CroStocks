import os
import openpyxl
import xlrd
import pandas as pd
from xlrd.sheet import ctype_text
import difflib
import csv
#POMOCNE FUNKCIJE
def list_equal(list):
    list=iter(list)
    try:
        first=next(list)
    except StopIteration:
        return True
    return all(first==rest for rest in list)




def sublist(list1,list2):
    ls=[item in list2 for item in list1]
    return all(ls)

def check_if_all_equal(temp):
    iterator=iter(temp)
    try:
        first=next(iterator)
    except StopIteration:
        return True
    return all(first==rest for rest in iterator)

def check_if_for_specific_values(temp):
    def find_element_in_column(elem, col, temp):
        for row_index in range(temp.nrows):
            if elem == temp.cell(row_index, col).value:
                break
        return row_index

def find_element_in_column(elem,col,temp):
    for row_index in range(temp.nrows):
        if elem==temp.cell(row_index,col).value:
            break
    return row_index


def find_element(value, temp):
    break_loop = False
    for row_index in range(temp.nrows):

        for col_index in range(temp.ncols):

            cell_object = str(temp.cell(row_index, col_index).value)

            if value in cell_object:
                break_loop = True
                break
        if break_loop:
            break
    return (row_index, col_index)


def print_row(row, temp):
    for col_index in range(temp.ncols):
        print(temp.cell(row, col_index).value)


def print_column(col, temp):
    for row_index in range(temp.nrows):
        print(temp.cell(row_index, col).value)
def find_element(value, temp):
    break_loop = False
    for row_index in range(temp.nrows):

        for col_index in range(temp.ncols):

            cell_object = str(temp.cell(row_index, col_index).value)

            if value in cell_object:
                break_loop = True
                break
        if break_loop:
            break
    return (row_index, col_index)


def print_row(row, temp):
    for col_index in range(temp.ncols):
        print(temp.cell(row, col_index).value)


def print_column(col, temp):
    for row_index in range(temp.nrows):
        print(temp.cell(row_index, col).value)


def find_element_in_column(elem,col,temp):
    for row_index in range(temp.nrows):
        if elem==str(temp.cell(row_index,col).value):
            break
    return row_index

# bilanca=pd.DataFrame()
# rdg=pd.DataFrame()
# nti=pd.DataFrame()
# ntd=pd.DataFrame()
# pk=pd.DataFrame()
# x=[]
# rez=[]
# temp=['Bilanca','RDG','NT_I']
# for item in os.listdir('D:\ZSE'):
#     curdir = os.path.join('D:\ZSE',item)
#     files=os.listdir(curdir)
#     for st in files:
#         if st[0:12]==(item+"-fin2018") and st[len(st)-4:]==".xls":
#             xl=xlrd.open_workbook(os.path.join(curdir,st))
#             x.append(xl.sheet_names())
#             rez.append(sublist(temp,xl.sheet_names()))
#             if(rez[-1]==0):
#                 print(item, st, sep="\t")
#
# xlfile=xlrd.open_workbook(os.path.join(curdir,st))
#
#
#
# #BILANCA
# temp_bilanca = xlfile.sheet_by_name("Bilanca")
# temp_rdg=xlfile.sheet_by_name("RDG")
# temp_nti=xlfile.sheet_by_name('NT_I')


def input_data(temp,data,aop_index,item):

        x,aop_columna=find_element('AOP',temp)

        begin_aop_index=find_element_in_column(aop_index,aop_columna,temp)

        text=[item]
        for i in range(begin_aop_index, temp.nrows):
            if str(temp.cell(i, aop_columna).value) != '':
                text.append(temp.cell(i,0).value)
        if data.index.size==0:
            data.loc[0]=text
        else:
            data.loc[data.index.size]=text
        return data


def findnth(string, substring, n):
    parts = string.split(substring, n + 1)
    if len(parts) <= n + 1:
        return -1
    return len(string) - len(parts[-1]) - len(substring)


def rows_equal(data):
    first=data.iloc[0]
    for i in range(1,data.shape[0]):
        print(i)
        if first.equals(data.iloc[i])==False:
            for j in range(1,data.iloc[i].shape[0]):
                print(j)
                if difflib.SequenceMatcher(None,first[j],data.iloc[i][j]).ratio()<0.85:
                    print('jest')
    return True



if __name__=='__main__':
    kreiranje = [False, False, False, False]
    pat=os.path.join(os.getcwd(),"Stock","functions",'BussinesCompanies.csv')
    nema =((pd.read_csv(pat,header=None)).iloc[0]).values.tolist()
    nema.append("HPB")
    nema.append('IKBA')
    nema.append('KABA')
    nema.append('KBZ')
    nema.append('PBZ')
    nema.append('PDBA')
    nema.append("ELPR")
    nema.append('TKPR')
    nema.append('SNBA')
    nema.append('ZABA')

    problem=["CROS","CROS2","ELPR","JDOS","TKPR"]
    for item in os.listdir('D:\ZSE'):
        if item in nema:
            continue
        else:
            category1.append(item)
        curdir = os.path.join('D:\ZSE', item)
        files = os.listdir(curdir)
        for st in files:
            if st[0:findnth(st, '-', 1)] == (item + "-fin2018") and (
                    st[len(st) - 4:] == ".xls" or st[len(st) - 5:] == ".xlsx"):
                xl = xlrd.open_workbook(os.path.join(curdir, st))
                print(item)

                # Bilanca

                if "Bilanca" in xl.sheet_names():
                    temp = xl.sheet_by_name('Bilanca')
                    x, aop_columna = find_element('AOP', temp)

                    # Kreiranja inicijalne bilance
                    if kreiranje[0] == False:
                        begin_aop_index = find_element_in_column("1.0", aop_columna, temp)
                        col_aop = ["Ticker"]
                        for i in range(begin_aop_index, temp.nrows):
                            if str(temp.cell(i, aop_columna).value) != '':
                                col_aop.append(str(temp.cell(i, aop_columna).value))
                        bilanca = pd.DataFrame(columns=col_aop)
                        kreiranje[0] = True

                    bilanca = input_data(temp, bilanca, '1.0',item)

                # RDG

                if "RDG" in xl.sheet_names():
                    temp = xl.sheet_by_name('RDG')
                    x, aop_columna = find_element('AOP', temp)

                    # Kreiranja inicijalnnog RDG-a
                    if kreiranje[1] == False:
                        begin_aop_index = find_element_in_column("111.0", aop_columna, temp)
                        col_aop = ["Ticker"]
                        for i in range(begin_aop_index, temp.nrows):
                            if str(temp.cell(i, aop_columna).value) != '':
                                col_aop.append(str(temp.cell(i, aop_columna).value))
                        rdg = pd.DataFrame(columns=col_aop)
                        kreiranje[1] = True

                    rdg = input_data(temp, rdg, '111.0',item)

                # NTI
                if "NT_I" in xl.sheet_names():
                    temp = xl.sheet_by_name('NT_I')
                    x, aop_columna = find_element('AOP', temp)

                    # Kreiranja inicijalne bilance
                    if kreiranje[2] == False:
                        begin_aop_index = find_element_in_column("1.0", aop_columna, temp)
                        col_aop = ["Ticker"]
                        for i in range(begin_aop_index, temp.nrows):
                            if str(temp.cell(i, aop_columna).value) != '':
                                col_aop.append(str(temp.cell(i, aop_columna).value))
                        nti = pd.DataFrame(columns=col_aop)
                        kreiranje[2] = True

                    nti = input_data(temp, nti, '1.0',item)

                # PROMJENA KAPITALA
                if "PK" in xl.sheet_names():
                    temp = xl.sheet_by_name('PK')
                    x, aop_columna = find_element('AOP', temp)

                    # Kreiranja inicijalne bilance
                    if kreiranje[3] == False:
                        begin_aop_index = find_element_in_column("1.0", aop_columna, temp)
                        col_aop = ["Ticker"]
                        for i in range(begin_aop_index, temp.nrows):
                            if str(temp.cell(i, aop_columna).value) != '':
                                col_aop.append(str(temp.cell(i, aop_columna).value))
                        pk = pd.DataFrame(columns=col_aop)
                        kreiranje[3] = True

                    pk = input_data(temp, pk, '1.0',item)

                break

    bilanca.to_excel("BilancaF.xlsx")
    rdg.to_excel("RDGF.xlsx")
    nti.to_excel("NTIF.xlsx")
    pk.to_excel("PKF.xlsx")

    # with open("BussinesCompanies.csv",'w') as f:
    #     wr=csv.writer(f)
    #     wr.writerow(category1)
    # f.close()
    #
    # with open("FinancialCompanies.csv",'w') as f:
    #     wr=csv.writer(f)
    #     wr.writerow(nema)
    # f.close()

    #PROVJERA