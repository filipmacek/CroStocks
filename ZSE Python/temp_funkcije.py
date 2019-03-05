import os
import openpyxl
import xlrd
import pandas as pd

bilanca=pd.DataFrame()
rdg=pd.DataFrame()
nti=pd.DataFrame()
ntd=pd.DataFrame()
pk=pd.DataFrame()
for item in os.listdir('D:\ZSE'):
    curdir=os.path.join('D:\ZSE',item)
    files=os.listdir(curdir)
    for st in files:
        if st[0:12]==(item+"-fin2018") and st[len(st)-4:]==".xls":
            print(item,st,sep="\t")
            break
xlfile=xlrd.open_workbook(os.path.join(curdir,st))

    #BILANCA
temp = xlfile.sheet_by_name("Bilanca")

num_cols=temp.ncols



for row_index in range(temp.nrows):
    print('-'*40)
    print("Row %s"%row_index)
    for col_index in range(temp.ncols):
        cell_object=temp.cell(row_index,col_index)
        print(cell_object)

def find_element(value, temp):
    break_loop = False
    for row_index in range(temp.nrows):

        for col_index in range(temp.ncols):

            cell_object = str(temp.cell(row_index, col_index).value)

            if value in cell_object:
                break_loop = True
                break
        if break_loop:
            break
    return (row_index, col_index)

x,y=find_element('AOP',temp)

for i in range(temp.nrows):
    print(temp.cell(i,y).value)
aop_indeks=x+3
ls=[]
for i in range(aop_indeks,temp.nrows):
    bilanca.loc[0]
